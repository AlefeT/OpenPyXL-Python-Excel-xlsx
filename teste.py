#!/usr/bin/python3

import openpyxl
from openpyxl import Workbook
import time


# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - #
                                #CRIA XLSX E POE DADOS NELE
#CRIA XLSX
#book = Workbook()
#sheet = book.active

#MODO 1 DE POR DADOS
#sheet['A1'] = 42

#MODO 2 DE POR DADOS
#sheet.cell(row=2, column=2).value = 3

#SALVA O XLSX
#book.save('sample.xlsx')

# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - #
                            #ABRE XLSX EXISTENTE E LE DADOS DELE

try:
    #CRIA O XLSX
    book = openpyxl.load_workbook('sample.xlsx')
    sheet = book.active

    #MODO 1 DE CAPTURAR DADOS
    #a1 = sheet['A1']
    #a2 = sheet['A2']
    #a3 = sheet.cell(row=3, column=1)

    #MOSTRA NA TELA
    #print(a1.value)
    #print(a2.value)
    #print(a3.value)



    #MODO 2 DE CAPTURAR DADOS
    #cells = sheet['A1': 'B6']

    #MOSTRA NA TELA
    #for c1, c2 in cells:
    #    print(c1.value, c2.value)



    #MODO 3 DE CAPTURAR DADOS JÁ MOSTRANDO NA TELA
    #for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    #    for cell in row:
    #        print(cell.value, end=" ")
    #    print()

        
    #MODO 4 DE CAPTURAR DADOS MOSTRANDO OS ROWS NA TELA        
    for row in sheet.values:
        print(row)



except Exception as E:
    print(E)

# - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - #


time.sleep(3)